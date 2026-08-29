#!/usr/bin/env python3
"""Extract casting-submission data from a Gmail Takeout .mbox.

Streams the archive, filters by Gmail label, reads each message for eleven
fields, saves genuine photos, and writes submissions.csv + audit.csv.

Standard library only (openpyxl optional, for --xlsx). See README.md.
"""
from __future__ import annotations

import argparse
import csv
import datetime as dt
import email
import email.policy
import email.utils
import hashlib
import html as htmllib
import mailbox
import re
import shutil
import struct
import sys
from collections import defaultdict
from pathlib import Path

try:
    from zoneinfo import ZoneInfo
    LONDON = ZoneInfo("Europe/London")
except Exception:                                    # pragma: no cover
    LONDON = dt.timezone.utc

HEADERS = [
    "Date Received", "First Name", "Last Name", "Sender Email Address",
    "Gender / Pronouns", "Age / Playing Age", "Phone Number", "Location",
    "Spotlight / Portfolio Link", "Drive Image Links", "Status",
]

# ---------------------------------------------------------------- constants

RELAY_DOMAINS = {
    "spotlight.com", "app.spotlight.com", "mandy.com", "backstage.com",
    "castingcallpro.com", "uk.castingcallpro.com", "starnow.com",
    "castingnetworks.com", "casting-networks.com", "tagmin.co.uk",
    "google.com", "docs.google.com", "wix.com", "squarespace.com",
    "formspree.io", "typeform.com", "jotform.com",
}
RELAY_LOCALS = {
    "no-reply", "noreply", "no_reply", "donotreply", "do-not-reply",
    "notifications", "notification", "mailer", "auto", "automated",
    "forms-receipts", "postmaster", "bounce",
}

HONORIFICS = {"mr", "mrs", "ms", "miss", "mx", "dr", "prof", "sir", "dame", "lady", "lord"}
SUFFIXES = {"jr", "sr", "ii", "iii", "iv", "phd", "ma", "ba"}
PARTICLES = {"van", "von", "de", "del", "della", "di", "du", "da", "das", "dos",
             "la", "le", "den", "der", "ten", "ter", "af", "av", "bin", "ibn", "al"}
ROLE_WORDS = {"actor", "actress", "performer", "artist", "dancer", "singer",
              "model", "spotlight", "equity", "voiceover", "vo", "he", "she", "they"}
NON_NAME_START = {
    "based", "interested", "looking", "writing", "attaching", "attached", "available",
    "keen", "excited", "happy", "currently", "originally", "free", "able", "just",
    "really", "very", "delighted", "reaching", "emailing", "applying", "submitting",
    "hoping", "also", "still", "not", "sorry", "sending", "from", "the", "and", "for",
    "with", "sure", "glad", "grateful", "back", "new", "here", "out", "about", "an", "a"}

SIGNOFF = (r"(?:kind regards|warm regards|best regards|many thanks|kind wishes|"
           r"best wishes|all the best|yours sincerely|yours faithfully|speak soon|"
           r"regards|sincerely|thanks so much|many thanks|thanks|thank you|cheers|warmly|best)")

LINK_RANK = [
    ("spotlight.com", 0), ("mandy.com", 1), ("backstage.com", 2),
    ("castingcallpro", 3), ("starnow.com", 4), ("imdb.com", 5),
    ("vimeo.com", 7), ("youtube.com", 7), ("youtu.be", 7),
    ("instagram.com", 8), ("linkedin.com", 9), ("facebook.com", 10),
]
LINK_BLOCK = re.compile(
    r"(?i)(unsubscribe|/privacy|/terms|/legal|/cookie|preferences|manage-?prefs|"
    r"googleusercontent|gstatic\.com|list-manage|mailchimp|sendgrid|mcusercontent|"
    r"doubleclick|/cdn-cgi/|\.(?:png|jpe?g|gif|webp|svg|css|js)(?:\?|$))")

JUNK_IMAGE_NAME = re.compile(
    r"(?i)(logo|icon|banner|signature|footer|spacer|pixel|beacon|social|award|badge|"
    r"twitter|facebook|linkedin|instagram|whatsapp|youtube|tiktok|crest|stamp)")
DOC_KEEP_NAME = re.compile(r"(?i)(headshot|head-shot|photo|portrait|cv|resume|r[ée]sum[ée]|spotlight|profile|casting)")

MIN_IMAGE_BYTES = 20 * 1024
MIN_IMAGE_EDGE = 300

# ------------------------------------------------------------- text helpers


def decode_part(part) -> str:
    try:
        return part.get_content()
    except Exception:
        raw = part.get_payload(decode=True) or b""
        cs = part.get_content_charset() or "utf-8"
        try:
            return raw.decode(cs, "replace")
        except LookupError:
            return raw.decode("utf-8", "replace")


def html_to_text(source: str):
    """Return (plain text, hrefs). Pull links out before the tags are stripped."""
    hrefs = re.findall(r"""(?i)href\s*=\s*["']([^"']+)["']""", source)
    t = re.sub(r"(?is)<(script|style)[^>]*>.*?</\1>", " ", source)
    t = re.sub(r"(?i)<br\s*/?>", "\n", t)
    t = re.sub(r"(?i)</(p|div|tr|li|h[1-6]|table)>", "\n", t)
    t = re.sub(r"<[^>]+>", " ", t)
    t = htmllib.unescape(t)
    t = t.replace("\xa0", " ")
    t = re.sub(r"[ \t]+", " ", t)
    t = re.sub(r"\n[ \t]+", "\n", t)
    t = re.sub(r"\n{3,}", "\n\n", t)
    return t.strip(), hrefs


def get_body(msg):
    """Best plain-text body plus every href found in any HTML alternative."""
    texts, htmls = [], []
    for part in msg.walk():
        if part.get_content_maintype() == "multipart":
            continue
        disp = str(part.get("Content-Disposition") or "").lower()
        if "attachment" in disp:
            continue
        ct = part.get_content_type()
        if ct == "text/plain":
            texts.append(decode_part(part))
        elif ct == "text/html":
            htmls.append(decode_part(part))
    hrefs = []
    for h in htmls:
        hrefs.extend(html_to_text(h)[1])
    plain = "\n".join(texts).strip()
    rich = "\n\n".join(html_to_text(h)[0] for h in htmls).strip()
    # Many senders ship a stub text/plain ("view this in HTML") next to the real
    # HTML body, so take whichever part actually carries the message.
    body = plain if len(plain) >= len(rich) else rich
    return body, hrefs


QUOTE_PATTERNS = [
    r"(?im)^[ \t]*On .{5,200}\bwrote:[ \t]*$",
    r"(?im)^[ \t]*-{2,}[ \t]*Original Message[ \t]*-{2,}",
    r"(?im)^[ \t]*_{20,}[ \t]*$",
    r"(?im)^[ \t]*From:[ \t]\S.*\n(?:.*\n){0,3}?[ \t]*(?:Sent|Date):[ \t]",
    r"(?im)^[ \t]*Begin forwarded message:",
    r"(?im)^[ \t]*-{3,}[ \t]*Forwarded message[ \t]*-{3,}",
]


def strip_quotes(body: str) -> str:
    """Drop quoted reply chains but keep the signature block.

    If cutting leaves almost nothing the message *is* the forwarded content
    (relay notifications look like this), so keep the whole thing.
    """
    cut = len(body)
    for pat in QUOTE_PATTERNS:
        m = re.search(pat, body)
        if m:
            cut = min(cut, m.start())
    head = body[:cut]
    head = "\n".join(ln for ln in head.splitlines() if not ln.lstrip().startswith(">"))
    head = head.strip()
    return head if len(head) >= 40 else body.strip()


def norm_ws(s: str) -> str:
    return re.sub(r"\s+", " ", s or "").strip()

# ------------------------------------------------------------ image helpers


def image_size(data: bytes):
    """(width, height) for PNG/JPEG/GIF/WEBP from the header bytes, else None."""
    try:
        if data[:8] == b"\x89PNG\r\n\x1a\n":
            w, h = struct.unpack(">II", data[16:24])
            return w, h
        if data[:3] == b"\xff\xd8\xff":
            i = 2
            while i < len(data) - 9:
                if data[i] != 0xFF:
                    i += 1
                    continue
                marker = data[i + 1]
                if marker in (0xD8, 0x01) or 0xD0 <= marker <= 0xD7:
                    i += 2
                    continue
                seglen = struct.unpack(">H", data[i + 2:i + 4])[0]
                if marker in (0xC0, 0xC1, 0xC2, 0xC3, 0xC5, 0xC6, 0xC7,
                              0xC9, 0xCA, 0xCB, 0xCD, 0xCE, 0xCF):
                    h, w = struct.unpack(">HH", data[i + 5:i + 9])
                    return w, h
                i += 2 + seglen
        if data[:6] in (b"GIF87a", b"GIF89a"):
            w, h = struct.unpack("<HH", data[6:10])
            return w, h
        if data[:4] == b"RIFF" and data[8:12] == b"WEBP" and data[12:16] == b"VP8 ":
            w, h = struct.unpack("<HH", data[26:30])
            return w & 0x3FFF, h & 0x3FFF
    except Exception:
        pass
    return None


def safe_filename(name: str, fallback: str = "file") -> str:
    name = (name or "").replace("\\", "/").split("/")[-1]
    name = re.sub(r"[^\w.\- ]+", "_", name).strip(" ._")
    return name or fallback

# --------------------------------------------------------- field extraction


class Found:
    __slots__ = ("value", "rule", "confidence")

    def __init__(self, value, rule, confidence="high"):
        self.value = value
        self.rule = rule
        self.confidence = confidence

    def __repr__(self):
        return f"Found({self.value!r}, {self.rule}, {self.confidence})"


def tidy_name(s: str) -> str:
    out = []
    toks = s.split()
    for i, t in enumerate(toks):
        if t.isupper() and len(t) > 1:
            t = t[0] + t[1:].lower()
            t = re.sub(r"^(O')(\w)", lambda m: m.group(1) + m.group(2).upper(), t)
            t = re.sub(r"^(Mac|Mc)(\w)", lambda m: m.group(1) + m.group(2).upper(), t)
        if t.lower() in PARTICLES and i < len(toks) - 1:
            t = t.lower()
        out.append(t)
    return " ".join(out)


def split_name(raw: str):
    """'SMITH, Jane' -> ('Jane','Smith'). Keeps particles and apostrophes intact."""
    s = norm_ws(raw)
    s = re.sub(r"\s*[\(\[].*?[\)\]]\s*", " ", s)
    s = re.split(r"\s+[\-–—|/]\s+", s)[0]
    s = re.sub(r'["“”]', "", s)
    s = norm_ws(s).strip(" ,.;:")
    if not s:
        return "", ""
    if "," in s:
        a, b = [p.strip() for p in s.split(",", 1)]
        b_toks = b.split()
        if (a and b_toks and len(a.split()) == 1 and len(b_toks) <= 2
                and not re.search(r"\d", b)
                and b_toks[0].lower().strip(".") not in ROLE_WORDS):
            return tidy_name(b), tidy_name(a)
        s = a
    toks = [t for t in s.split() if t.lower().strip(".") not in HONORIFICS]
    while toks and toks[-1].lower().strip(".") in SUFFIXES:
        toks.pop()
    if not toks:
        return "", ""
    if len(toks) == 1:
        return tidy_name(toks[0]), ""
    i = len(toks) - 1
    while i > 1 and toks[i - 1].lower().strip(".") in PARTICLES:
        i -= 1
    return tidy_name(" ".join(toks[:i])), tidy_name(" ".join(toks[i:]))


def looks_like_person(s: str) -> bool:
    s = norm_ws(s)
    if not s or "@" in s or re.search(r"\d{3}", s):
        return False
    if re.search(r"(?i)\b(ltd|limited|agency|management|casting|productions?|studios?|"
                 r"team|admin|talent|associates|group|llp|inc)\b", s):
        return False
    toks = s.split()
    if not toks or toks[0].lower().strip(".,") in NON_NAME_START:
        return False
    return 1 <= len(toks) <= 4


def titlecase_name(s):
    return " ".join(t[:1].upper() + t[1:] for t in s.split())


def find_name(body, display_name, addr, attach_names, subject=""):
    """Resolve a name, then top it up with a surname if the display name has one."""
    got = _find_name_core(body, display_name, addr, attach_names, subject)
    if got and len(got.value.split()) == 1 and looks_like_person(display_name or ""):
        d = norm_ws(display_name)
        if len(d.split()) >= 2 and d.split()[0].lower().strip(",") == got.value.split()[0].lower():
            return Found(d, got.rule + "+display-name-surname", got.confidence)
    return got


def _find_name_core(body, display_name, addr, attach_names, subject=""):
    # No (?i) on anything using [A-Z] - the flag would cancel the capital guard
    # and let "I'm based in Hackney" parse as a name.
    for pat, rule in [
        (r"(?im)^[ \t]*(?:full[ \t]+)?(?:name|applicant|candidate|actor|performer)"
         r"[ \t]*[:\-][ \t]*(.{2,60})$", "explicit-name-field"),
        (r"[Mm]y name(?:'s| is)\s+([^\n,.!]{2,60})", "explicit-statement"),
        (r"\b[Tt]his is\s+([A-Z][\w'\-]+(?:\s+[A-Z][\w'\-]+){1,2})\b", "explicit-statement"),
    ]:
        m = re.search(pat, body)
        if m and looks_like_person(m.group(1)):
            return Found(m.group(1).strip(), rule, "high")

    # An agency submitting a client: the row belongs to the actor, not the agent.
    m = re.search(r"\b(?:[Ss]ubmit(?:ting)?|[Pp]utting forward|[Rr]epresent(?:ing)?|"
                  r"[Pp]roposing)\s+([A-Z][\w'\-]+(?:\s+[A-Z][\w'\-]+){1,2})\b",
                  subject + "\n" + body)
    if m and looks_like_person(m.group(1)):
        return Found(m.group(1).strip(), "agent-submission", "high")

    m = re.search(r"\bI(?:'m| am)\s+([A-Z][\w'\-]+(?:\s+[A-Z][\w'\-]+){1,2})\b", body)
    if m and looks_like_person(m.group(1)):
        return Found(m.group(1).strip(), "explicit-statement", "high")

    m = re.search(rf"(?im)^[ \t]*{SIGNOFF}[,!.]*[ \t]*\n+[ \t]*([^\n]{{2,60}})[ \t]*$", body)
    if m and looks_like_person(m.group(1)):
        return Found(m.group(1).strip(), "signature-block", "high")

    if looks_like_person(display_name):
        return Found(display_name.strip(), "from-display-name", "high")

    for fn in attach_names:
        stem = re.sub(r"\.[A-Za-z0-9]{2,5}$", "", fn)
        stem = re.sub(r"(?i)[_\-\s]*(headshot|head-shot|photo|portrait|cv|resume|"
                      r"r[ée]sum[ée]|spotlight|profile|casting|final|\d+)[_\-\s]*", " ", stem)
        stem = norm_ws(stem.replace("_", " ").replace("-", " "))
        if looks_like_person(stem) and len(stem.split()) >= 2:
            return Found(titlecase_name(stem), "attachment-filename", "medium")

    local = addr.split("@")[0]
    if re.fullmatch(r"[a-z]+[._\-][a-z]+\d{0,3}", local, re.I):
        parts = re.split(r"[._\-]", re.sub(r"\d+$", "", local))
        return Found(" ".join(p.capitalize() for p in parts if p), "email-local-part", "low")
    return None


def find_first_name_only(body):
    m = re.search(r"\bI(?:'m| am)\s+([A-Z][a-z'\-]{1,20})\b(?=[,.!\n]|\s+and\b|\s+an?\b)", body)
    if m and m.group(1).lower() not in ROLE_WORDS:
        return Found(m.group(1), "explicit-statement-firstname", "medium")
    return None


PRONOUN_RE = re.compile(
    r"\b(she|he|they|ze|xe|ey|per|fae)\s*/\s*(her|him|them|hir|xem|em|per|faer|they|she|he)"
    r"(?:\s*/\s*(hers|his|theirs|hirs|xyrs))?\b", re.I)
GENDER_FIELD = re.compile(r"(?im)^[ \t]*(?:gender|sex)[ \t]*[:\-][ \t]*([A-Za-z \-/]{3,30})[ \t]*$")
SELF_DESC = re.compile(
    r"(?i)\b(?:I(?:'m| am)|as)\s+an?\s+(?:(trans|cis)\s+)?"
    r"(man|woman|male|female|non[-\s]?binary|nonbinary|actress)\b")


def find_pronouns(body):
    """Explicit statements only. Never inferred from a name or a photo."""
    pronouns = gender = None
    rule = []
    m = PRONOUN_RE.search(body)
    if m:
        pronouns = "/".join(g.lower() for g in m.groups() if g)
        rule.append("pronoun-string")
    m = GENDER_FIELD.search(body)
    if m:
        gender = norm_ws(m.group(1)).title()
        rule.append("gender-field")
    if not gender:
        m = SELF_DESC.search(body)
        if m:
            word = m.group(2).lower()
            prefix = (m.group(1) or "").lower()
            base = {"man": "Male", "male": "Male", "woman": "Female",
                    "female": "Female", "actress": "Female"}.get(word, "Non-binary")
            gender = f"{prefix.capitalize()} {base.lower()}".strip() if prefix else base
            rule.append("self-description-actress" if word == "actress" else "self-description")
    if gender and pronouns:
        return Found(f"{gender} — {pronouns}", "+".join(rule), "high")
    if pronouns:
        return Found(pronouns, "+".join(rule), "high")
    if gender:
        conf = "medium" if "actress" in "".join(rule) else "high"
        return Found(gender, "+".join(rule), conf)
    return None


PLAYING_PATS = [
    (r"(?i)playing\s*age\s*(?:range)?\s*[:\-]?\s*(\d{1,2})\s*(?:[-–—]|to)\s*(\d{1,2})", "playing-range"),
    (r"(?i)\b(?:I\s+)?plays?\s+(?:from\s+)?(\d{1,2})\s*(?:[-–—]|to)\s*(\d{1,2})\b", "playing-range"),
    (r"(?i)playing\s*age\s*[:\-]?\s*(\d{1,2})\b", "playing-single"),
]
ACTUAL_PATS = [
    (r"(?im)^[ \t]*age[ \t]*[:\-][ \t]*(\d{1,2})\b", "age-field"),
    (r"(?i)\bI'?\s?a?m\s+[A-Z][\w'\-]+,\s*(\d{1,2})\b", "age-statement"),
    (r"(?i)\bI'?\s?a?m\s+(\d{1,2})\s*(?:years?\s*old)?\b(?!\s*(?:[-–—]|to)\s*\d)", "age-statement"),
    (r"(?i)\b(\d{1,2})\s*years?\s*old\b", "age-statement"),
]
DOB_PAT = re.compile(
    r"(?i)\b(?:d\.?o\.?b\.?|date of birth|born)\b\s*(?:on|is|was|:|-|=)?\s*"
    r"(\d{1,2}[/.\-]\d{1,2}[/.\-]\d{2,4}"
    r"|\d{1,2}(?:st|nd|rd|th)?\s+[A-Za-z]{3,9}\s+\d{4}"
    r"|[A-Za-z]{3,9}\s+\d{1,2}(?:st|nd|rd|th)?,?\s+\d{4}"
    r"|\d{4}-\d{2}-\d{2})")


def parse_dob(s):
    s = re.sub(r"(?i)(\d)(?:st|nd|rd|th)", r"\1", s.strip().rstrip(".,;")).replace(",", "")
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y", "%Y-%m-%d", "%d %B %Y",
                "%d %b %Y", "%B %d %Y", "%b %d %Y", "%d/%m/%y"):
        try:
            return dt.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def find_age(body, received):
    """Playing age beats actual age beats DOB. A graduation year is not an age."""
    for pat, rule in PLAYING_PATS:
        m = re.search(pat, body)
        if m:
            gs = [g for g in m.groups() if g]
            if len(gs) == 2:
                lo, hi = int(gs[0]), int(gs[1])
                if 3 <= lo < hi <= 99:
                    return Found(f"Playing {lo}–{hi}", rule, "high")
            elif gs and 3 <= int(gs[0]) <= 99:
                return Found(f"Playing {int(gs[0])}", rule, "high")

    scrub = re.sub(r"(?i)playing\s*age[^\n]*", " ", body)
    for pat, rule in ACTUAL_PATS:
        m = re.search(pat, scrub)
        if m:
            n = int(m.group(1))
            if 5 <= n <= 99:
                return Found(str(n), rule, "high")

    m = DOB_PAT.search(body)
    if m:
        d = parse_dob(m.group(1))
        if d and received:
            age = received.year - d.year - ((received.month, received.day) < (d.month, d.day))
            if 5 <= age <= 99:
                return Found(f"{age} (from DOB)", "derived-from-dob", "high")
    return None

PHONE_CAND = re.compile(r"(?<![\w\d])(\+?\d[\d\s().\-]{7,20}\d)(?![\w\d])")
BAD_PHONE_CTX = re.compile(
    r"(?i)(vat|company\s*(?:no|number|reg)|registered\s*(?:in|no|office)|invoice|"
    r"sort\s*code|account\s*(?:no|number)|charity|passport|policy|postcode|"
    r"unsubscribe|ref(?:erence)?\s*(?:no|number|:)|pin\b)")


def normalise_phone(raw: str):
    d = re.sub(r"[^\d+]", "", raw)
    if d.startswith("00"):
        d = "+" + d[2:]
    elif d.startswith("0"):
        d = "+44" + d[1:]
    elif not d.startswith("+"):
        if len(d) == 10 and d.startswith("7"):
            d = "+44" + d
        else:
            return None, False
    body = d[1:]
    if not (9 <= len(body) <= 15):
        return None, False
    if body.startswith("44"):
        rest = body[2:]
        if len(rest) != 10:
            return None, False
        if rest.startswith("7"):
            return f"+44 {rest[:4]} {rest[4:]}", True
        if rest[0] == "2":
            return f"+44 {rest[:2]} {rest[2:6]} {rest[6:]}", False
        return f"+44 {rest[:3]} {rest[3:6]} {rest[6:]}", False
    spaced = re.sub(r"\s+", " ", raw.strip())
    if not spaced.startswith("+"):
        spaced = "+" + body
    return spaced, False


def find_phone(body):
    """Prefer a mobile, and the number nearest the sign-off."""
    best = None
    for m in PHONE_CAND.finditer(body):
        ls = body.rfind("\n", 0, m.start()) + 1
        le = body.find("\n", m.end())
        line = body[ls: le if le != -1 else len(body)]
        if BAD_PHONE_CTX.search(line):
            continue
        raw = m.group(1)
        if len(re.sub(r"\D", "", raw)) < 9:
            continue
        norm, mobile = normalise_phone(raw)
        if not norm:
            continue
        agent = bool(re.search(r"(?i)(agent|agency|management|office|switchboard|studio)", line))
        score = (2 if mobile else 0) + (0 if agent else 1)
        cand = (score, m.start(), norm, mobile, agent)
        if best is None or (cand[0], cand[1]) > (best[0], best[1]):
            best = cand
    if not best:
        return None, []
    others = []
    for m in PHONE_CAND.finditer(body):
        n, _ = normalise_phone(m.group(1))
        if n and n != best[2] and n not in others:
            others.append(n)
    rule = "signature-mobile" if best[3] else ("agent-number" if best[4] else "body-number")
    return Found(best[2], rule, "high" if best[3] else "medium"), others


LOC_PATS = [
    (r"(?im)^[ \t]*(?:location|based|city|town|area)[ \t]*[:\-][ \t]*([^\n]{2,40})$", "location-field"),
    (r"\b(?:[Bb]ased|[Ll]iving|[Ll]ocated)\s+(?:in|near|just outside|just north of|"
     r"just south of|around|close to|outside)\s+([A-Z][\w'’\-]*(?:[ ][A-Z][\w'’\-]*){0,2})",
     "explicit-statement"),
    (r"\bI(?:'m| am)\s+(?:currently\s+)?(?:based|living)\s+in\s+"
     r"([A-Z][\w'’\-]*(?:[ ][A-Z][\w'’\-]*){0,2})", "explicit-statement"),
    (r"\b(?:I|[Ss]he|[Hh]e|[Tt]hey)\s+(?:currently\s+)?(?:live|lives|reside|resides)\s+"
     r"(?:in|near)\s+([A-Z][\w'’\-]*(?:[ ][A-Z][\w'’\-]*){0,2})", "explicit-statement"),
    (r"\b([A-Z][a-z]+(?:[ -][A-Z][a-z]+)?)[- ]based\b", "explicit-statement"),
]
POSTCODE = re.compile(r"\b([A-Z]{1,2}\d{1,2}[A-Z]?)\s*\d[A-Z]{2}\b")
OUTWARD = {
    "M": "Manchester", "B": "Birmingham", "LS": "Leeds", "L": "Liverpool",
    "G": "Glasgow", "EH": "Edinburgh", "CF": "Cardiff", "BS": "Bristol",
    "NE": "Newcastle upon Tyne", "S": "Sheffield", "NG": "Nottingham",
    "BN": "Brighton", "OX": "Oxford", "CB": "Cambridge", "LE": "Leicester",
    "CV": "Coventry", "BA": "Bath", "YO": "York", "RG": "Reading",
}
LONDON_OUT = {"E", "EC", "N", "NW", "SE", "SW", "W", "WC"}
LONDON_AREAS = {
    "hackney", "peckham", "camden", "islington", "brixton", "shoreditch", "clapham",
    "greenwich", "walthamstow", "stratford", "croydon", "ealing", "hammersmith",
    "dalston", "deptford", "lewisham", "tottenham", "wimbledon", "fulham",
    "kilburn", "brockley", "bermondsey", "hendon", "soho", "dulwich",
}
UK_CITIES = {
    "london", "manchester", "birmingham", "leeds", "liverpool", "glasgow",
    "edinburgh", "cardiff", "bristol", "newcastle", "sheffield", "nottingham",
    "brighton", "oxford", "cambridge", "leicester", "coventry", "bath", "york",
    "reading", "salford", "belfast", "aberdeen", "southampton", "norwich",
}
MULTIWORD_UK = {"newcastle upon tyne": "Newcastle upon Tyne",
                "stoke on trent": "Stoke-on-Trent",
                "milton keynes": "Milton Keynes",
                "stratford upon avon": "Stratford-upon-Avon"}
NON_UK = {"dublin": "Ireland", "cork": "Ireland", "galway": "Ireland",
          "new york": "USA", "los angeles": "USA", "toronto": "Canada",
          "sydney": "Australia", "berlin": "Germany", "paris": "France",
          "amsterdam": "Netherlands", "madrid": "Spain"}


LOC_JOIN = re.compile(r"(?:\s+(?:upon|on|under|by)\s+[A-Z][\w'’\-]+)")
LOC_TRAIL = re.compile(r"(?i)\s+(?:but|and|so|though|although|where|while|with|for|"
                       r"because|however|as|until|during)\b.*$")


def canon_location(place: str, rule: str, conf: str = "high"):
    p = LOC_TRAIL.sub("", norm_ws(place)).strip(" ,.;:").rstrip(".")
    if not p:
        return None
    low = p.lower()
    if low in LONDON_AREAS:
        return Found("London, UK", rule + "+london-area", conf)
    if low in NON_UK:
        return Found(f"{p.title()}, {NON_UK[low]}", rule, conf)
    if low in MULTIWORD_UK:
        return Found(f"{MULTIWORD_UK[low]}, UK", rule, conf)
    for city in UK_CITIES:
        if low == city or low.startswith(city + " "):
            return Found(f"{city.title()}, UK", rule, conf)
    return Found(p, rule, "medium")


def find_location(body):
    for pat, rule in LOC_PATS:
        m = re.search(pat, body)
        if m:
            place = m.group(1)
            tail = LOC_JOIN.match(body, m.end(1))
            if tail:
                place += tail.group(0)
            got = canon_location(place, rule)
            if got:
                return got
    m = POSTCODE.search(body)
    if m:
        out = m.group(1)
        area = re.match(r"[A-Z]{1,2}", out).group(0)
        if area in LONDON_OUT:
            return Found("London, UK", "postcode-lookup", "medium")
        if area in OUTWARD:
            return Found(f"{OUTWARD[area]}, UK", "postcode-lookup", "medium")
    return None


URL_RE = re.compile(r"""(?i)\b(?:https?://|www\.)[^\s<>"')\]}]+""")


def clean_url(u: str) -> str:
    u = u.strip().rstrip(".,;:)]}>\"'")
    u = re.sub(r"(?i)([?&])(utm_[^=]+|fbclid|gclid|mc_cid|mc_eid)=[^&]*", r"\1", u)
    u = re.sub(r"[?&]+$", "", u).replace("?&", "?")
    if u.lower().startswith("www."):
        u = "https://" + u
    return u


def rank_url(u: str) -> int:
    low = u.lower()
    for frag, r in LINK_RANK:
        if frag in low:
            return r
    return 6


def find_links(body, hrefs):
    seen, out = set(), []
    for raw in list(URL_RE.findall(body)) + list(hrefs):
        if raw.lower().startswith("mailto:"):
            continue
        u = clean_url(raw)
        if not u.lower().startswith("http") or LINK_BLOCK.search(u):
            continue
        key = u.lower().rstrip("/")
        if key in seen:
            continue
        seen.add(key)
        out.append(u)
    if not out:
        return None
    out.sort(key=lambda u: (rank_url(u), len(u)))
    rule = "spotlight-link" if rank_url(out[0]) == 0 else "portfolio-link"
    return Found(" | ".join(out), rule, "high" if rank_url(out[0]) <= 5 else "medium")


def is_relay(addr: str) -> bool:
    if "@" not in addr:
        return False
    local, domain = addr.rsplit("@", 1)
    return domain.lower() in RELAY_DOMAINS or local.lower() in RELAY_LOCALS


def find_sender(msg, body):
    from_addr = (email.utils.parseaddr(str(msg.get("From") or ""))[1] or "").lower()
    if not is_relay(from_addr):
        return Found(from_addr, "from-header", "high") if from_addr else None
    reply = (email.utils.parseaddr(str(msg.get("Reply-To") or ""))[1] or "").lower()
    if reply and not is_relay(reply):
        return Found(reply, "reply-to-header", "high")
    m = re.search(r"(?im)^[ \t]*e-?mail[ \t]*[:\-][ \t]*([\w.+\-]+@[\w.\-]+\.\w+)", body)
    if m and not is_relay(m.group(1).lower()):
        return Found(m.group(1).lower(), "body-email-field", "high")
    for cand in re.findall(r"[\w.+\-]+@[\w.\-]+\.\w{2,}", body):
        if not is_relay(cand.lower()) and cand.lower() != from_addr:
            return Found(cand.lower(), "body-email-scan", "medium")
    return Found(from_addr, "relay-only", "low") if from_addr else None

# ------------------------------------------------------------------- media


def collect_media(msg, media_log, msg_id):
    """Return [(filename, bytes, sha)] of parts that look like real photos."""
    kept, hashes = [], set()
    for part in msg.walk():
        if part.get_content_maintype() == "multipart":
            continue
        ct = part.get_content_type()
        name = safe_filename(part.get_filename() or "", "")
        is_img = ct.startswith("image/")
        is_doc = ct == "application/pdf"
        if not (is_img or is_doc):
            continue
        try:
            data = part.get_payload(decode=True) or b""
        except Exception:
            media_log.append((msg_id, name or ct, ct, 0, "", "discard", "undecodable"))
            continue
        size = len(data)
        sha = hashlib.sha256(data).hexdigest()
        dims = image_size(data) if is_img else None
        dim_s = f"{dims[0]}x{dims[1]}" if dims else ""

        reason = None
        if is_doc and not DOC_KEEP_NAME.search(name):
            reason = "pdf-not-headshot-or-cv"
        elif JUNK_IMAGE_NAME.search(name):
            reason = "junk-filename"
        elif size < MIN_IMAGE_BYTES:
            reason = f"under-{MIN_IMAGE_BYTES // 1024}kb"
        elif dims and min(dims) < MIN_IMAGE_EDGE:
            reason = f"under-{MIN_IMAGE_EDGE}px"
        elif sha in hashes:
            reason = "duplicate-in-message"

        if reason:
            media_log.append((msg_id, name, ct, size, dim_s, "discard", reason))
            continue
        hashes.add(sha)
        ext = Path(name).suffix or {"image/jpeg": ".jpg", "image/png": ".png",
                                    "image/heic": ".heic", "image/webp": ".webp",
                                    "application/pdf": ".pdf"}.get(ct, ".bin")
        media_log.append((msg_id, name, ct, size, dim_s, "keep", ""))
        kept.append((name or f"attachment{ext}", data, sha, ct, size, dim_s))
    return kept


# --------------------------------------------------------------- per-message


def process_message(msg, idx, media_log):
    raw_from = str(msg.get("From") or "")
    display = email.utils.parseaddr(raw_from)[0]
    msg_id = str(msg.get("Message-ID") or f"<no-id-{idx}>")

    received = None
    try:
        d = email.utils.parsedate_to_datetime(str(msg.get("Date")))
        if d:
            if d.tzinfo is None:
                d = d.replace(tzinfo=dt.timezone.utc)
            received = d.astimezone(LONDON)
    except Exception:
        pass

    body_raw, hrefs = get_body(msg)
    body = strip_quotes(body_raw)
    media = collect_media(msg, media_log, msg_id)
    attach_names = [m[0] for m in media] + [
        safe_filename(p.get_filename() or "", "") for p in msg.walk() if p.get_filename()]

    fields, notes = {}, []
    fields["Date Received"] = (Found(received.strftime("%Y-%m-%d %H:%M"), "date-header", "high")
                               if received else None)

    subject = str(msg.get("Subject") or "")
    name = find_name(body, display, email.utils.parseaddr(raw_from)[1] or "",
                     attach_names, subject)
    if name:
        first, last = split_name(name.value)
        fields["First Name"] = Found(first, name.rule, name.confidence) if first else None
        fields["Last Name"] = Found(last, name.rule, name.confidence) if last else None
    else:
        fields["First Name"] = fields["Last Name"] = None
    if not fields.get("First Name"):
        fo = find_first_name_only(body)
        if fo:
            fields["First Name"] = fo

    fields["Sender Email Address"] = find_sender(msg, body_raw)
    fields["Gender / Pronouns"] = find_pronouns(body)
    fields["Age / Playing Age"] = find_age(body, received.date() if received else None)
    phone, other_phones = find_phone(body)
    fields["Phone Number"] = phone
    if other_phones:
        notes.append("other numbers seen: " + ", ".join(other_phones))
    fields["Location"] = find_location(body)
    fields["Spotlight / Portfolio Link"] = find_links(body, hrefs)
    fields["Drive Image Links"] = None

    for pat, label in [(r"(?i)willing to travel", "willing to travel"),
                       (r"(?i)self[- ]tape", "self-tape mentioned"),
                       (r"(?i)\bgraduat(?:ed|ing|e)\b", "graduation mentioned (not used as age)"),
                       (r"(?i)\brelocat", "relocating")]:
        if re.search(pat, body):
            notes.append(label)

    age = fields.get("Age / Playing Age")
    if age and age.rule.startswith("playing"):
        scrub = re.sub(r"(?i)playing\s*age[^\n]*", " ", body)
        for pat, _ in ACTUAL_PATS:
            m2 = re.search(pat, scrub)
            if m2 and 5 <= int(m2.group(1)) <= 99:
                notes.append(f"actual age also stated: {m2.group(1)}")
                break

    return {
        "idx": idx, "msg_id": msg_id, "from": raw_from, "subject": subject,
        "labels": str(msg.get("X-Gmail-Labels") or ""), "received": received,
        "fields": fields, "media": media, "notes": "; ".join(notes),
    }


# ------------------------------------------------------------------- merging


def merge_records(records):
    groups, order = {}, []
    for r in records:
        se = r["fields"].get("Sender Email Address")
        key = se.value if se and se.value else f"__idx{r['idx']}"
        if key not in groups:
            groups[key] = []
            order.append(key)
        groups[key].append(r)

    # second pass: same name + a shared phone or link means a second address
    merged_into = {}
    keys = list(order)
    for i, a in enumerate(keys):
        if a in merged_into:
            continue
        for b in keys[i + 1:]:
            if b in merged_into:
                continue
            fa, fb = groups[a][0]["fields"], groups[b][0]["fields"]

            def v(f, k):
                return (f.get(k).value if f.get(k) else "")
            if not (v(fa, "First Name") and v(fa, "Last Name")):
                continue
            if (v(fa, "First Name"), v(fa, "Last Name")) != (v(fb, "First Name"), v(fb, "Last Name")):
                continue
            if v(fa, "Phone Number") and v(fa, "Phone Number") == v(fb, "Phone Number"):
                merged_into[b] = a
            elif v(fa, "Spotlight / Portfolio Link") and v(fa, "Spotlight / Portfolio Link") == v(fb, "Spotlight / Portfolio Link"):
                merged_into[b] = a
    for b, a in merged_into.items():
        groups[a].extend(groups[b])
        order.remove(b)

    out = []
    for key in order:
        rs = sorted(groups[key], key=lambda r: (r["received"] or dt.datetime.max.replace(tzinfo=LONDON)))
        row, prov, conflicts = {}, {}, []
        for h in HEADERS:
            if h in ("Drive Image Links", "Status"):
                continue
            chosen = None
            for r in rs:
                f = r["fields"].get(h)
                if f and f.value:
                    if chosen is None:
                        chosen = f
                    elif f.value != chosen.value:
                        conflicts.append(f"{h}: kept '{chosen.value}', also saw '{f.value}'")
            row[h] = chosen.value if chosen else ""
            prov[h] = f"{chosen.rule}/{chosen.confidence}" if chosen else "not-found"
        if rs[0]["received"]:
            row["Date Received"] = rs[0]["received"].strftime("%Y-%m-%d %H:%M")
            prov["Date Received"] = "date-header/high"
        media = []
        seen = set()
        for r in rs:
            for m in r["media"]:
                if m[2] not in seen:
                    seen.add(m[2])
                    media.append(m)
        notes = "; ".join(n for n in [r["notes"] for r in rs] if n)
        if conflicts:
            notes = (notes + "; " if notes else "") + "conflicts -> " + "; ".join(conflicts)
        if len(rs) > 1:
            notes = (notes + "; " if notes else "") + f"merged {len(rs)} emails"
        out.append({"row": row, "prov": prov, "media": media, "sources": rs, "notes": notes})
    return out


def compute_status(row, media):
    have_name = bool(row["First Name"])
    have_email = bool(row["Sender Email Address"])
    contactable = any(row[k] for k in ("Phone Number", "Location", "Spotlight / Portfolio Link"))
    if not (have_name and have_email and contactable):
        return "Needs review"
    if not media:
        return "No media"
    return "New"


def slug_for(row, idx):
    first, last = row["First Name"], row["Last Name"]
    if first or last:
        s = f"{last}_{first}".strip("_")
    else:
        addr = row["Sender Email Address"] or f"unknown{idx}"
        s = "unknown_" + addr.split("@")[0]
    return re.sub(r"[^\w\-]+", "_", s).strip("_") or f"unknown_{idx}"

# ------------------------------------------------------------------ outputs


def write_xlsx(path, rows):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
    except ImportError:
        return False
    wb = Workbook()
    ws = wb.active
    ws.title = "Submissions"
    ws.append(HEADERS)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(vertical="center")
    for r in rows:
        ws.append([r["row"].get(h, "") for h in HEADERS])
    for col, h in enumerate(HEADERS, start=1):
        width = max([len(h)] + [len(str(r["row"].get(h, ""))) for r in rows]) + 2
        ws.column_dimensions[ws.cell(1, col).column_letter].width = min(max(width, 12), 48)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(path)
    return True


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--mbox", required=True)
    ap.add_argument("--out", default="./casting-submissions")
    ap.add_argument("--label", help="X-Gmail-Labels value to filter on (substring, case-insensitive)")
    ap.add_argument("--list-labels", action="store_true", help="print the label histogram and exit")
    ap.add_argument("--limit", type=int, help="process only the first N matching messages (dry run)")
    ap.add_argument("--no-media", action="store_true")
    ap.add_argument("--xlsx", action="store_true")
    args = ap.parse_args()

    mbox_path = Path(args.mbox).expanduser()
    if not mbox_path.exists():
        sys.exit(f"mbox not found: {mbox_path}")
    out = Path(args.out).expanduser()
    out.mkdir(parents=True, exist_ok=True)

    mb = mailbox.mbox(str(mbox_path),
                      factory=lambda f: email.message_from_binary_file(f, policy=email.policy.default),
                      create=False)

    if args.list_labels:
        hist = defaultdict(int)
        for msg in mb:
            for lab in str(msg.get("X-Gmail-Labels") or "(none)").split(","):
                hist[lab.strip() or "(none)"] += 1
        print(f"{len(mb)} messages\n")
        for lab, n in sorted(hist.items(), key=lambda kv: -kv[1]):
            print(f"{n:>6}  {lab}")
        return

    media_log, records = [], []
    total = 0
    for i, msg in enumerate(mb):
        total += 1
        if args.label:
            labs = str(msg.get("X-Gmail-Labels") or "")
            if args.label.lower() not in labs.lower():
                continue
        records.append(process_message(msg, i, media_log))
        if args.limit and len(records) >= args.limit:
            break

    print(f"scanned {total} messages, selected {len(records)}")
    merged = merge_records(records)

    media_root = out / "media"
    flat = media_root / "_all-images"
    if not args.no_media:
        media_root.mkdir(exist_ok=True)
        flat.mkdir(exist_ok=True)

    for n, rec in enumerate(merged, 1):
        slug = slug_for(rec["row"], n)
        rec["slug"] = slug
        paths = []
        if not args.no_media and rec["media"]:
            d = media_root / slug
            d.mkdir(parents=True, exist_ok=True)
            for k, (name, data, sha, ct, size, dims) in enumerate(rec["media"], 1):
                ext = Path(name).suffix.lower() or ".bin"
                fn = f"{slug}_{k:02d}{ext}"
                (d / fn).write_bytes(data)
                shutil.copyfile(d / fn, flat / fn)
                paths.append(str((d / fn).resolve()))
        rec["paths"] = paths
        rec["row"]["Drive Image Links"] = " | ".join(paths)
        rec["row"]["Status"] = compute_status(rec["row"], rec["media"])
        rec["prov"]["Drive Image Links"] = "local-path" if paths else "not-found"
        rec["prov"]["Status"] = "derived"

    csv_path = out / "submissions.csv"
    with open(csv_path, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(HEADERS)
        for rec in merged:
            w.writerow([rec["row"].get(h, "") for h in HEADERS])

    audit_path = out / "audit.csv"
    with open(audit_path, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["Applicant", "Message-ID", "From header", "Subject", "Gmail labels"]
                   + [f"{h} :: value" for h in HEADERS[:-2]]
                   + [f"{h} :: rule" for h in HEADERS[:-2]] + ["Images kept", "Notes"])
        for rec in merged:
            for src in rec["sources"]:
                f = src["fields"]
                w.writerow(
                    [rec["slug"], src["msg_id"], src["from"], src["subject"], src["labels"]]
                    + [(f.get(h).value if f.get(h) else "") for h in HEADERS[:-2]]
                    + [(f"{f.get(h).rule}/{f.get(h).confidence}" if f.get(h) else "not-found")
                       for h in HEADERS[:-2]]
                    + [len(src["media"]), rec["notes"]])

    with open(out / "media-log.csv", "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["Message-ID", "Filename", "MIME", "Bytes", "Dimensions", "Decision", "Reason"])
        w.writerows(media_log)

    if args.xlsx:
        if write_xlsx(out / "submissions.xlsx", merged):
            print(f"wrote {out / 'submissions.xlsx'}")
        else:
            print("openpyxl not installed - skipped xlsx (pip install openpyxl)")

    # ------------------------------------------------------------- summary
    print(f"wrote {csv_path}\nwrote {audit_path}\n")
    print(f"{len(merged)} applicant rows from {len(records)} emails")
    print("\nfill rate:")
    for h in HEADERS:
        n = sum(1 for r in merged if r["row"].get(h))
        print(f"  {h:<32} {n:>3}/{len(merged)}  {100 * n / max(len(merged), 1):>5.0f}%")
    print("\nstatus:")
    st = defaultdict(int)
    for r in merged:
        st[r["row"]["Status"]] += 1
    for k, v in sorted(st.items(), key=lambda kv: -kv[1]):
        print(f"  {k:<32} {v}")
    low = [(r["slug"], h, r["prov"][h]) for r in merged for h in HEADERS if r["prov"].get(h, "").endswith("/low")]
    if low:
        print("\nlow-confidence extractions (check these):")
        for slug, h, rule in low:
            print(f"  {slug:<28} {h:<30} {rule}")
    print(f"\nimages kept {sum(len(r['media']) for r in merged)}, "
          f"discarded {sum(1 for m in media_log if m[5] == 'discard')} "
          f"(see media-log.csv)")


if __name__ == "__main__":
    main()
